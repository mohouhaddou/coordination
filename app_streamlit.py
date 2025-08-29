# app_streamlit.py
# -*- coding: utf-8 -*-
# Analyse de brouillage (CHIRplus_BC) par seuil ENU + exports Excel/PDF + carte interactive.
# - Parsing largeur fixe robuste (y compris rapports sans table)
# - Réparation chiffres tronqués (ex. DIS "07 4.1" -> "074.1")
# - Nettoyage TRANSMITTER (suppression des chiffres collés)
# - Colonnes Canal/Fréquence fiables (en-tête -> table -> nom de fichier; normalisation UHF)
# - Carte pydeck (fonds CARTO/Mapbox), affichage du SITE étudié (Longit./Latit.)
# - Filtrage combiné "Site + Canal" (sélecteurs)
# - 2e tableau Interférences filtré sur (Site, Canal) sélectionnés
# - Export Excel sécurisé (nettoyage caractères XML illégaux)

import io
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional

import pandas as pd
import streamlit as st
import pydeck as pdk

import json
import xml.etree.ElementTree as ET
import zipfile
import base64
from pydeck.types import String as PdkString


from pandas.io.formats.style import Styler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # ✅ pour nettoyer XML
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

st.set_page_config(page_title="Analyse Brouillage ENU (CHIRplus_BC)", layout="wide")

# =============================
# Colonnes numériques à réparer (espaces internes)
# =============================
NUMERIC_COLUMNS = {'ENU', 'DIS', 'AZM', 'ERP', 'f/MHz', 'CHA', 'HEFF'}

def _compact_internal_spaces_in_number(s: str) -> str:
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    return re.sub(r'(?<=\d)\s+(?=\d)', '', s)

def _repair_numeric_cells(row: Dict[str, Any]) -> Dict[str, Any]:
    for k in NUMERIC_COLUMNS:
        if k in row and row[k]:
            row[k] = _compact_internal_spaces_in_number(row[k])
    return row

# =============================
# Sanitation Excel (anti IO_WRITE)
# =============================
def sanitize_cell(x):
    if isinstance(x, str):
        return ILLEGAL_CHARACTERS_RE.sub("", x)
    return x

def sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    return df.applymap(sanitize_cell)

# =============================
# Parsing helpers
# =============================
def _find_table_lines(lines: List[str]) -> Tuple[int, int]:
    header_idx = -1
    pattern = re.compile(r'^\s*ENU\s+OS\s+TRANSMITTER', re.IGNORECASE)
    for i, ln in enumerate(lines):
        if pattern.search(ln):
            header_idx = i
            break
    if header_idx == -1:
        return -1, -1
    start = header_idx
    end = len(lines)
    for j in range(header_idx+1, len(lines)):
        if lines[j].strip() == "" and (j+1 < len(lines) and lines[j+1].strip() == ""):
            end = j
            break
    return start, end

def _column_slices(header: str):
    starts = []
    h = header.rstrip("\n")
    for i, ch in enumerate(h):
        if (i == 0 and ch != " ") or (i > 0 and ch != " " and h[i-1] == " "):
            starts.append(i)
    ends = starts[1:] + [len(h)]
    cols = []
    for s, e in zip(starts, ends):
        name = h[s:e].strip()
        if name:
            cols.append((name, s, e))
    return cols

def _parse_fixed_width_row(line: str, cols):
    row = {}
    for name, s, e in cols:
        row[name] = line[s:e].strip()
    return row

def parse_table_from_txt(txt: str) -> List[Dict[str, Any]]:
    lines = txt.splitlines()
    start, end = _find_table_lines(lines)
    if start == -1:
        return []
    header = lines[start]
    cols = _column_slices(header)
    rows = []
    for ln in lines[start+1:end]:
        if not ln.strip():
            continue
        if not re.match(r'^\s*\d', ln):
            continue
        r = _parse_fixed_width_row(ln, cols)
        r = _repair_numeric_cells(r)
        rows.append(r)
    return rows

def extract_site_name(txt: str, default_name: str) -> str:
    patterns = [
        r'Interf\.\s*Transmit\.?\s*:\s*(.+)',
        r'Interfer\.\s*Transmit\.?\s*:\s*(.+)',
        r'Interf.*Transmit.*:\s*(.+)',
    ]
    for pat in patterns:
        m = re.search(pat, txt, flags=re.IGNORECASE)
        if m:
            name = m.group(1).strip().splitlines()[0].strip()
            return name
    m = re.search(r'Filename\s*:\s*(.+)', txt, flags=re.IGNORECASE)
    if m:
        candidate = Path(m.group(1).strip()).stem
        if candidate:
            return candidate
    return default_name

# =============================
# KML/KMZ → GeoJSON (avancé)
# =============================
def _parse_kml_coordinates(coord_text: str) -> list:
    coords = []
    if not coord_text:
        return coords
    for token in coord_text.replace("\n", " ").split():
        parts = token.split(",")
        if len(parts) >= 2:
            try:
                lon = float(parts[0]); lat = float(parts[1])
                coords.append([lon, lat])
            except Exception:
                pass
    return coords

def _feature_point(props, coords_text):
    coords = _parse_kml_coordinates(coords_text)
    if coords:
        return {"type": "Feature", "properties": props, "geometry": {"type": "Point", "coordinates": coords[0]}}
    return None

def _feature_linestring(props, coords_text):
    coords = _parse_kml_coordinates(coords_text)
    if len(coords) >= 2:
        return {"type": "Feature", "properties": props, "geometry": {"type": "LineString", "coordinates": coords}}
    return None

def _feature_polygon(props, coords_text):
    ring = _parse_kml_coordinates(coords_text)
    if len(ring) >= 3:
        if ring[0] != ring[-1]:
            ring.append(ring[0])
        return {"type": "Feature", "properties": props, "geometry": {"type": "Polygon", "coordinates": [ring]}}
    return None

def _extract_props(pm, ns):
    props = {}
    name_el = pm.find("k:name", ns)
    desc_el = pm.find("k:description", ns)
    if name_el is not None and name_el.text:
        props["name"] = name_el.text
    if desc_el is not None and desc_el.text:
        props["description"] = desc_el.text
    return props

def kml_kmz_to_geojson(file_name: str, file_bytes: bytes) -> Optional[dict]:
    """
    GeoJSON FeatureCollection depuis KML/KMZ
    - Placemark: Point/LineString/Polygon
    - MultiGeometry
    - Document/Folder
    """
    if file_name.lower().endswith(".kmz"):
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            kml_name = None
            for n in zf.namelist():
                if n.lower().endswith("doc.kml"):
                    kml_name = n; break
            if not kml_name:
                kml_name = next((n for n in zf.namelist() if n.lower().endswith(".kml")), None)
            if not kml_name:
                return None
            raw = zf.read(kml_name)
            try:
                kml_text = raw.decode("utf-8")
            except UnicodeDecodeError:
                kml_text = raw.decode("latin-1", errors="ignore")
    else:
        try:
            kml_text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            kml_text = file_bytes.decode("latin-1", errors="ignore")

    try:
        root = ET.fromstring(kml_text)
    except Exception:
        return None

    ns = {"k": "http://www.opengis.net/kml/2.2"}
    features = []

    for pm in root.findall(".//k:Placemark", ns):
        props = _extract_props(pm, ns)
        mg = pm.find(".//k:MultiGeometry", ns)
        if mg is not None:
            for c in mg.findall(".//k:Point/k:coordinates", ns):
                f = _feature_point(props, c.text or "");   features.append(f) if f else None
            for c in mg.findall(".//k:LineString/k:coordinates", ns):
                f = _feature_linestring(props, c.text or ""); features.append(f) if f else None
            for c in mg.findall(".//k:Polygon/k:outerBoundaryIs/k:LinearRing/k:coordinates", ns):
                f = _feature_polygon(props, c.text or "");    features.append(f) if f else None
            continue

        cpt = pm.find(".//k:Point/k:coordinates", ns)
        if cpt is not None and cpt.text:
            f = _feature_point(props, cpt.text); features.append(f) if f else None
            continue

        cls = pm.find(".//k:LineString/k:coordinates", ns)
        if cls is not None and cls.text:
            f = _feature_linestring(props, cls.text); features.append(f) if f else None
            continue

        cpg = pm.find(".//k:Polygon/k:outerBoundaryIs/k:LinearRing/k:coordinates", ns)
        if cpg is not None and cpg.text:
            f = _feature_polygon(props, cpg.text); features.append(f) if f else None
            continue

    if not features:
        return None
    return {"type": "FeatureCollection", "features": features}

# =============================
# KMZ GroundOverlay (image → BitmapLayer)
# =============================
def _read_text_guess_enc(raw: bytes) -> str:
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("latin-1", errors="ignore")

def _build_data_url(image_bytes: bytes, href: str) -> str:
    href_lower = href.lower()
    if href_lower.endswith(".jpg") or href_lower.endswith(".jpeg"):
        mime = "image/jpeg"
    elif href_lower.endswith(".webp"):
        mime = "image/webp"
    else:
        mime = "image/png"
    import base64 as _b64
    b64 = _b64.b64encode(image_bytes).decode("ascii")
    return f"data:{mime};base64,{b64}"

def _overlay_from_go_element(go_el: ET.Element, image_lookup: callable) -> Optional[dict]:
    name = None
    name_el = go_el.find(".//{*}name")
    if name_el is not None and name_el.text:
        name = name_el.text.strip()

    llb = go_el.find(".//{*}LatLonBox")
    if llb is None:
        return None
    try:
        north = float((llb.find(".//{*}north").text or "").strip())
        south = float((llb.find(".//{*}south").text or "").strip())
        east  = float((llb.find(".//{*}east").text  or "").strip())
        west  = float((llb.find(".//{*}west").text  or "").strip())
    except Exception:
        return None
    bounds = [west, south, east, north]

    href_el = go_el.find(".//{*}Icon/{*}href")
    if href_el is None or not href_el.text:
        return None
    href = href_el.text.strip()

    img_bytes = image_lookup(href)
    if not img_bytes:
        return None
    data_url = _build_data_url(img_bytes, href)
    return {"name": name, "image_data_url": data_url, "bounds": bounds}

def extract_groundoverlays_from_kml(kml_text: str, asset_loader: callable) -> list:
    try:
        root = ET.fromstring(kml_text)
    except Exception:
        return []
    overlays = []
    for go in root.findall(".//{*}GroundOverlay"):
        ov = _overlay_from_go_element(go, asset_loader)
        if ov:
            overlays.append(ov)
    return overlays

def kmz_extract_groundoverlays(file_name: str, file_bytes: bytes) -> list:
    overlays = []
    if file_name.lower().endswith(".kmz"):
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            kml_name = None
            for n in zf.namelist():
                if n.lower().endswith("doc.kml"):
                    kml_name = n; break
            if not kml_name:
                kml_name = next((n for n in zf.namelist() if n.lower().endswith(".kml")), None)
            if not kml_name:
                return []
            kml_text = _read_text_guess_enc(zf.read(kml_name))

            def _asset_loader(href: str) -> Optional[bytes]:
                cand = href
                if cand not in zf.namelist():
                    base_dir = str(Path(kml_name).parent).replace("\\", "/")
                    if base_dir == ".": base_dir = ""
                    alt = f"{base_dir}/{href}" if base_dir else href
                    cand = alt
                try:
                    return zf.read(cand)
                except KeyError:
                    return None

            overlays = extract_groundoverlays_from_kml(kml_text, _asset_loader)
    else:
        kml_text = _read_text_guess_enc(file_bytes)
        def _no_image(_): return None
        overlays = extract_groundoverlays_from_kml(kml_text, _no_image)
        overlays = [ov for ov in overlays if ov.get("image_data_url")]
    return overlays

# ========= Extraction fiable (en-tête) =========
def extract_header_freq_channel(txt: str) -> Tuple[Optional[float], Optional[int]]:
    freq = None; ch = None
    m_f = re.search(r'Frequency\s*/?\s*MHz\s*[:=]\s*([0-9]+(?:\.[0-9]+)?)', txt, flags=re.IGNORECASE)
    if m_f:
        try: freq = float(m_f.group(1))
        except Exception: freq = None
    m_c = re.search(r'Chan(?:nel)?\s*[:=]\s*(\d+)', txt, flags=re.IGNORECASE)
    if m_c:
        try: ch = int(m_c.group(1))
        except Exception: ch = None
    return freq, ch

# ========= Extraction coordonnées du site (en-tête) =========
def extract_site_coords(txt: str) -> Tuple[Optional[str], Optional[str]]:
    m = re.search(r'Longit\.\s*/\s*Latit\.\s*:\s*([0-9NSEW\s]+)\s*/\s*([0-9NSEW\s]+)', txt, flags=re.IGNORECASE)
    if not m: return None, None
    lon = m.group(1).strip(); lat = m.group(2).strip()
    return lon, lat

# ========= Fallback depuis nom de fichier =========
def infer_freq_channel_from_filename(name: str) -> Tuple[Optional[float], Optional[int]]:
    m = re.search(r'(\d+(?:\.\d+)?)', name)
    if not m: return None, None
    try: f = float(m.group(1))
    except Exception: return None, None
    ch = None
    if 470.0 <= f <= 862.0:
        ch = int(round((f - 474.0)/8.0) + 21)
    return f, ch

# ========= Normalisation fréquence depuis canal (UHF 8 MHz) =========
def freq_from_channel_uhf(ch: int) -> Optional[float]:
    if 21 <= ch <= 69:
        return 306.0 + 8.0*ch
    return None

# ========= Nettoyage nom d'émetteur =========
def clean_transmitter_name(name: Optional[str]) -> Optional[str]:
    if not name: return name
    return re.sub(r'\d+$', '', str(name)).strip()

# =============================
# Filtrage au-dessus du seuil
# =============================
def rows_above_threshold(rows: List[Dict[str, Any]], threshold: float) -> List[Dict[str, Any]]:
    out = []
    for r in rows:
        enu_str = r.get('ENU', '')
        try:
            enu = float(enu_str)
        except Exception:
            mm = re.findall(r'[-+]?\d+(?:\.\d+)?', enu_str)
            if not mm: continue
            enu = float(mm[0])
        if enu > threshold:
            r = dict(r)
            r['_ENU_float'] = enu
            if 'TRANSMITTER' in r:
                r['TRANSMITTER'] = clean_transmitter_name(r['TRANSMITTER'])
            out.append(r)
    out.sort(key=lambda x: x['_ENU_float'], reverse=True)
    return out

def _safe_int(x) -> Optional[int]:
    try:
        return int(re.findall(r'\d+', str(x))[0])
    except Exception:
        return None

def process_one_text(content: str, filename: str, threshold: float):
    site = extract_site_name(content, default_name=Path(filename).stem)
    freq_h, ch_h = extract_header_freq_channel(content)
    lon_str, lat_str = extract_site_coords(content)
    lon_dd = None; lat_dd = None

    table_rows = parse_table_from_txt(content)
    interfs = rows_above_threshold(table_rows, threshold) if table_rows else []

    canal = ch_h; freq = freq_h
    if table_rows:
        canal = canal if canal is not None else _safe_int(table_rows[0].get('CHA'))
        val_f = table_rows[0].get('f/MHz')
        if freq is None and val_f:
            mf = re.search(r'([0-9]{2,4}(?:\.[0-9]+)?)', val_f)
            if mf:
                try: freq = float(mf.group(1))
                except Exception: pass

    if freq is None or canal is None:
        f_guess, ch_guess = infer_freq_channel_from_filename(filename)
        if freq is None and f_guess is not None: freq = f_guess
        if canal is None and ch_guess is not None: canal = ch_guess

    if (freq is None or (isinstance(freq, (int, float)) and freq < 100.0)) and (canal is not None):
        f_norm = freq_from_channel_uhf(canal)
        if f_norm is not None: freq = f_norm

    if lon_str and lat_str:
        lon_dd = None if lon_str is None else dms_to_decimal(lon_str)
        lat_dd = None if lat_str is None else dms_to_decimal(lat_str)

    risk = len(interfs) > 0
    worst = interfs[0] if risk else None
    summary = {
        'site': site,
        'Canal': canal,
        'Fréquence': freq,
        'file': filename,
        'threshold_ENU_dB': threshold,
        'risk': 'OUI' if risk else 'NON',
        'interferer_count': len(interfs),
        'max_ENU': worst['_ENU_float'] if worst else None,
        'worst_transmitter': clean_transmitter_name(worst['TRANSMITTER']) if worst and 'TRANSMITTER' in worst else None,
        'site_lon': lon_dd,
        'site_lat': lat_dd,
    }
    for r in interfs:
        r['site'] = site
        r['Canal'] = canal
        r['Fréquence'] = freq
        r['file'] = filename
        if 'TRANSMITTER' in r:
            r['TRANSMITTER'] = clean_transmitter_name(r['TRANSMITTER'])
    return summary, interfs

# =============================
# Couleurs
# =============================
def color_scale_enu(enu: float, threshold: float) -> str:
    if enu <= threshold: return "#E6F4EA"
    diff = min(max(enu - threshold, 0.0), 20.0) / 20.0
    if diff < 0.33: return "#FFF3C4"
    elif diff < 0.66: return "#FFD08A"
    else: return "#F8B4B4"

def hex_to_rgb(hex_color: str, alpha: int = 190) -> list:
    hex_color = hex_color.lstrip("#")
    return [int(hex_color[i:i+2], 16) for i in (0, 2, 4)] + [alpha]

def style_summary(df: pd.DataFrame, threshold: float) -> Styler:
    def highlight_risk(val):
        if val == 'OUI': return 'background-color: #F8B4B4; font-weight: 600;'
        elif val == 'NON': return 'background-color: #E6F4EA; font-weight: 600;'
        return ''
    def color_max_enu(v):
        if pd.isna(v): return ''
        return f'background-color: {color_scale_enu(float(v), threshold)}'
    return df.style.applymap(highlight_risk, subset=['risk']).applymap(color_max_enu, subset=['max_ENU'])

def style_details(df: pd.DataFrame, threshold: float) -> Styler:
    def color_enu(v):
        if pd.isna(v): return ''
        return f'background-color: {color_scale_enu(float(v), threshold)}'
    subset_cols = [c for c in ['_ENU_float','ENU'] if c in df.columns]
    return df.style.applymap(color_enu, subset=subset_cols)

# =============================
# DMS → degrés décimaux
# =============================
def dms_to_decimal(s: str) -> Optional[float]:
    if not s: return None
    s = s.strip()
    m = re.search(r'([0-9]{1,3})\s*([NSEW])\s*([0-9]{1,2})(?:\s+([0-9]{1,2}))?', s, flags=re.IGNORECASE)
    if not m: return None
    deg = int(m.group(1)); hemi = m.group(2).upper()
    minutes = int(m.group(3)); seconds = int(m.group(4)) if m.group(4) else 0
    val = deg + minutes/60.0 + seconds/3600.0
    if hemi in ('S', 'W'): val = -val
    return val

# =============================
# Styles de carte
# =============================
CARTO_LIGHT = "https://basemaps.cartocdn.com/gl/positron-gl-style/style.json"
CARTO_DARK  = "https://basemaps.cartocdn.com/gl/dark-matter-gl-style/style.json"
MAPBOX_LIGHT = "mapbox://styles/mapbox/light-v9"
MAPBOX_SATELLITE = "mapbox://styles/mapbox/satellite-v9"
# Remplacez l’existante par celle-ci
ESRI_WORLD_IMAGERY = (
    "https://server.arcgisonline.com/ArcGIS/rest/services/"
    "World_Imagery/MapServer/tile/{z}/{y}/{x}.jpg"
)

GOOGLE_SAT = "https://tile.opentopomap.org/{z}/{x}/{y}.png"


def get_map_style(style_name: str, mapbox_token: Optional[str]) -> Tuple[Optional[str], bool]:
    if style_name == "Clair (CARTO)":
        return CARTO_LIGHT, False
    if style_name == "Sombre (CARTO)":
        return CARTO_DARK, False
    if style_name == "Satellite (ESRI – gratuit)":
        return "", True   # ← style vide + on ajoutera la TileLayer ESRI
    if style_name == "Satellite (Google – sans clé)":
        return "", True  # on ajoute TileLayer manuellement

    if style_name == "Satellite (Mapbox)":
        if mapbox_token:
            pdk.settings.mapbox_api_key = mapbox_token
            return MAPBOX_SATELLITE, False
        return CARTO_LIGHT, False
    if style_name == "Mapbox Light":
        if mapbox_token:
            pdk.settings.mapbox_api_key = mapbox_token
            return MAPBOX_LIGHT, False
        return CARTO_LIGHT, False
    return CARTO_LIGHT, False



# =============================
# Excel export (mise en forme)
# =============================
def export_to_excel(df_sum: pd.DataFrame, df_det: pd.DataFrame) -> bytes:
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    desired = ['site','Canal','Fréquence','file','threshold_ENU_dB','risk','interferer_count','max_ENU','worst_transmitter','site_lon','site_lat']
    df_sum2 = df_sum[[c for c in desired if c in df_sum.columns] + [c for c in df_sum.columns if c not in desired]].copy()
    df_sum2 = sanitize_df_for_excel(df_sum2)

    ws1 = wb.active; ws1.title = "Résumé"
    for r in dataframe_to_rows(df_sum2, index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=ws1.max_column):
        for cell in row: cell.border = border
        headers = [c.value for c in ws1[1]]
        if 'risk' in headers:
            idx = headers.index('risk') + 1
            val = row[idx-1].value
            row[idx-1].fill = PatternFill("solid", fgColor=("F8B4B4" if val == 'OUI' else "E6F4EA"))
        if 'max_ENU' in headers and not df_sum.empty:
            j = headers.index('max_ENU') + 1
            v = row[j-1].value
            if v is not None:
                thr = float(df_sum['threshold_ENU_dB'].iloc[0])
                row[j-1].fill = PatternFill("solid", fgColor=color_scale_enu(float(v), thr).replace('#',''))

    ws2 = wb.create_sheet("Interférences")
    if df_det is None or df_det.empty:
        df_det = pd.DataFrame(columns=['site','Canal','Fréquence','file','TRANSMITTER','ENU'])
    else:
        preferred = ['site','Canal','Fréquence','file','ENU','OS','TRANSMITTER','DIS','AZM','LONGITUDE','LATITUDE','ERP','f/MHz','CHA','HEFF','POL','PROGRAM','REMARKS','_ENU_float']
        df_det = df_det[[c for c in preferred if c in df_det.columns] + [c for c in df_det.columns if c not in preferred]].copy()
    df_det = sanitize_df_for_excel(df_det)

    for r in dataframe_to_rows(df_det, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    headers2 = [c.value for c in ws2[1]]
    enu_cols = [headers2.index(k)+1 for k in ['_ENU_float','ENU'] if k in headers2]
    thr = float(df_sum['threshold_ENU_dB'].iloc[0]) if not df_sum.empty else 0.0
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=ws2.max_column):
        for cell in row: cell.border = border
        for j in enu_cols:
            v = row[j-1].value
            if v is None: continue
            try:
                color = color_scale_enu(float(v), thr).replace('#','')
                row[j-1].fill = PatternFill("solid", fgColor=color)
            except Exception:
                pass

    for ws in [ws1, ws2]:
        for col in ws.columns:
            maxlen = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(maxlen+2, 50)

    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()

# =============================
# PDF export
# =============================
def export_to_pdf(df_sum: pd.DataFrame, df_det: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Analyse de Brouillage (ENU) – Synthèse", styles['Title']))
    story.append(Spacer(1, 8))

    if not df_sum.empty:
        desired = ['site','Canal','Fréquence','file','threshold_ENU_dB','risk','interferer_count','max_ENU','worst_transmitter','site_lon','site_lat']
        df_sum2 = df_sum[[c for c in desired if c in df_sum.columns] + [c for c in df_sum.columns if c not in desired]].copy()
    else:
        df_sum2 = df_sum

    if df_sum2.empty:
        story.append(Paragraph("Aucune donnée disponible.", styles['Normal']))
    else:
        data_sum = [list(df_sum2.columns)] + df_sum2.astype(str).values.tolist()
        t1 = Table(data_sum, repeatRows=1)
        t1.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ]))
        headers = list(df_sum2.columns)
        if 'risk' in headers:
            c = headers.index('risk')
            for i, val in enumerate(df_sum2['risk'].tolist(), start=1):
                bg = colors.HexColor('#F8B4B4') if val=='OUI' else colors.HexColor('#E6F4EA')
                t1.setStyle(TableStyle([('BACKGROUND', (c,i), (c,i), bg)]))
        if 'max_ENU' in headers:
            j = headers.index('max_ENU')
            thr = float(df_sum2['threshold_ENU_dB'].iloc[0])
            for i, val in enumerate(df_sum2['max_ENU'].tolist(), start=1):
                if pd.isna(val): continue
                bg = colors.HexColor(color_scale_enu(float(val), thr))
                t1.setStyle(TableStyle([('BACKGROUND', (j,i), (j,i), bg)]))
        story.append(t1)

    story.append(Spacer(1, 16))
    story.append(Paragraph("Interférences (ENU > seuil) – Détail", styles['Heading2']))

    if df_det is not None and not df_det.empty:
        preferred = ['site','Canal','Fréquence','file','TRANSMITTER','ENU','_ENU_float','DIS','AZM','LONGITUDE','LATITUDE','ERP','f/MHz','CHA','HEFF','POL','PROGRAM','REMARKS']
        df_det2 = df_det[[c for c in preferred if c in df_det.columns] + [c for c in df_det.columns if c not in preferred]].copy()
        data_det = [list(df_det2.columns)] + df_det2.astype(str).values.tolist()
        t2 = Table(data_det, repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ]))
        headers2 = list(df_det2.columns)
        thr = float(df_sum['threshold_ENU_dB'].iloc[0]) if not df_sum.empty else 0.0
        for key in ['_ENU_float','ENU']:
            if key in headers2:
                idx = headers2.index(key)
                for i, v in enumerate(df_det2[key].tolist(), start=1):
                    try:
                        bg = colors.HexColor(color_scale_enu(float(v), thr))
                        t2.setStyle(TableStyle([('BACKGROUND', (idx,i), (idx,i), bg)]))
                    except Exception:
                        pass
        story.append(t2)
    else:
        story.append(Paragraph("Aucune interférence au-dessus du seuil.", styles['Normal']))

    doc.build(story)
    return buffer.getvalue()

# =============================
# Carte interactive (pydeck)
# =============================
def build_map_dataframe(df: pd.DataFrame, threshold: float, only_interferers: bool,
                        marker_base: float, marker_scale: float) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    enu_col = '_ENU_float' if '_ENU_float' in df.columns else 'ENU'
    df2 = df.copy()
    if only_interferers and enu_col in df2.columns:
        df2 = df2[pd.to_numeric(df2[enu_col], errors='coerce') > threshold]

    def _lon(x): return dms_to_decimal(x) if pd.notna(x) else None
    def _lat(x): return dms_to_decimal(x) if pd.notna(x) else None
    if 'LONGITUDE' in df2.columns and 'LATITUDE' in df2.columns:
        df2['lon'] = df2['LONGITUDE'].apply(_lon)
        df2['lat'] = df2['LATITUDE'].apply(_lat)
        df2 = df2.dropna(subset=['lon','lat'])
    else:
        return pd.DataFrame()

    def _color(row):
        try: val = float(row[enu_col])
        except Exception: return hex_to_rgb("#9E9E9E", 120)
        return hex_to_rgb(color_scale_enu(val, threshold), 190)

    def _radius(row):
        try: val = float(row[enu_col])
        except Exception: val = threshold
        gain = max(0.0, min(val - threshold, 20.0))
        return float(marker_base) + float(marker_scale) * gain

    df2['color_rgba'] = df2.apply(_color, axis=1)
    df2['radius_m']   = df2.apply(_radius, axis=1)

    for c in ['site','Canal','Fréquence','TRANSMITTER','OS','DIS','AZM','ERP','PROGRAM','REMARKS','file','f/MHz','CHA','HEFF','ENU']:
        if c not in df2.columns:
            df2[c] = None
    if 'ENU' not in df2.columns and enu_col in df2.columns:
        df2['ENU'] = df2[enu_col]
    return df2

def build_sites_dataframe(df_sum: pd.DataFrame) -> pd.DataFrame:
    if df_sum is None or df_sum.empty or ('site_lon' not in df_sum.columns) or ('site_lat' not in df_sum.columns):
        return pd.DataFrame()
    df_sites = df_sum[['site','Canal','Fréquence','site_lon','site_lat']].dropna()
    if df_sites.empty: return df_sites
    df_sites = df_sites.rename(columns={'site_lon':'lon','site_lat':'lat'})
    df_sites['TRANSMITTER'] = '— (site étudié)'
    df_sites['ENU'] = None; df_sites['DIS'] = None; df_sites['AZM'] = None; df_sites['ERP'] = None
    df_sites['PROGRAM'] = None; df_sites['REMARKS'] = None; df_sites['file'] = None
    return df_sites

def render_map(df_emitters: pd.DataFrame, df_sites: pd.DataFrame, title: str = "Carte des émetteurs",
               map_style_choice: str = "Clair (CARTO)", mapbox_token: Optional[str] = None,
               site_marker_radius: float = 2000.0,
               geojson_overlay: Optional[dict] = None,
               ground_overlays: Optional[list] = None,
               overlay_opacity: float = 0.65,
               show_overlay: bool = True):

    if (df_emitters is None or df_emitters.empty) and (df_sites is None or df_sites.empty) \
       and not geojson_overlay and not (show_overlay and ground_overlays):
        st.info("Aucun point/couche à afficher sur la carte.")
        return

    lats, lons = [], []
    if df_emitters is not None and not df_emitters.empty:
        lats += df_emitters['lat'].tolist(); lons += df_emitters['lon'].tolist()
    if df_sites is not None and not df_sites.empty:
        lats += df_sites['lat'].tolist();    lons += df_sites['lon'].tolist()

    if (not lats or not lons) and (geojson_overlay and geojson_overlay.get("features")):
        try:
            def _collect_coords(g):
                t = g.get("type"); c = g.get("coordinates")
                if t == "Point": return [c]
                if t == "LineString": return c
                if t == "Polygon": return c[0] if c else []
                return []
            pts = []
            for f in geojson_overlay.get("features", []):
                pts += _collect_coords(f.get("geometry", {}))
            if pts:
                lons = [p[0] for p in pts if isinstance(p, (list, tuple)) and len(p) >= 2]
                lats = [p[1] for p in pts if isinstance(p, (list, tuple)) and len(p) >= 2]
        except Exception:
            pass

    mid_lat = float(pd.Series(lats).median()) if lats else 31.7
    mid_lon = float(pd.Series(lons).median()) if lons else -7.1

    layers = []
       # === Choix du fond
    map_style, use_esri_tile = get_map_style(map_style_choice, mapbox_token)

    layers = []

    # --- (A) FOND ESRI : TileLayer en première position ---
    if use_esri_tile:
        layers.append(pdk.Layer(
            "TileLayer",
            data={"url": GOOGLE_SAT},   # ← Google satellite
            minZoom=0,
            maxZoom=19,
            tileSize=256,
            opacity=1.0
    ))


    # --- (B) ÉMETTEURS ---
    if df_emitters is not None and not df_emitters.empty:
        layers.append(pdk.Layer(
            "ScatterplotLayer",
            df_emitters,
            get_position='[lon, lat]',
            get_radius='radius_m',
            get_fill_color='color_rgba',
            pickable=True,
            stroked=True,
            get_line_color=[60, 60, 60],
            line_width_min_pixels=1,
        ))

    # --- (C) SITES ---
    if df_sites is not None and not df_sites.empty:
        layers.append(pdk.Layer(
            "ScatterplotLayer",
            df_sites,
            get_position='[lon, lat]',
            get_radius=site_marker_radius,
            get_fill_color=[30,144,255,200],
            pickable=True,
            stroked=True,
            get_line_color=[0,0,0],
            line_width_min_pixels=1,
        ))

    # --- (D) Overlays (GeoJSON / Bitmap) inchangés ---
    if geojson_overlay:
        layers.append(pdk.Layer(
            "GeoJsonLayer",
            geojson_overlay,
            pickable=True,
            stroked=True,
            filled=True,
            get_fill_color=[255,255,0,60],
            get_line_color=[255,0,0,200],
            line_width_min_pixels=2,
            get_point_radius=6,
            point_radius_min_pixels=4,
        ))
    if show_overlay and ground_overlays:
        for ov in ground_overlays:
            layers.append(pdk.Layer(
                "BitmapLayer",
                data=None,
                image=ov["image_data_url"],
                bounds=ov["bounds"],   # [west, south, east, north]
                opacity=float(overlay_opacity),
                desaturate=0.0
            ))

    tooltip = {...}
    view_state = pdk.ViewState(latitude=mid_lat, longitude=mid_lon, zoom=6, min_zoom=2, pitch=0)

    st.subheader(title)

    # === Construction du Deck ===
    if use_esri_tile:
        deck = pdk.Deck(
            layers=layers,
            initial_view_state=view_state,
            tooltip=tooltip,
            map_style=""   # ← très important en pydeck 0.9.1
        )
    else:
        deck = pdk.Deck(
            layers=layers,
            initial_view_state=view_state,
            tooltip=tooltip,
            map_style=map_style
        )

    st.pydeck_chart(deck)



# =============================
# UI
# =============================
st.title("Analyse de Brouillage par ENU – Rapports CHIRplus_BC")
st.caption("Chargez un ou plusieurs rapports TXT, choisissez un seuil ENU, filtrez (Site+Canal), visualisez la carte et exportez en Excel/PDF.")

with st.sidebar:
    st.header("Paramètres généraux")
    threshold = st.number_input("Seuil ENU (dB)", min_value=-200.0, max_value=200.0, value=60.0, step=1.0, format="%.1f")

    st.header("Carte – Affichage")
    only_map_interferers = st.checkbox("N'afficher que les émetteurs > seuil", value=True)
    marker_base = st.slider("Taille de base des émetteurs (m)", 200, 8000, 800, step=100)
    marker_scale = st.slider("Agrandissement par dB (m/dB)", 20, 800, 120, step=10)
    site_marker_radius = st.slider("Taille du marker SITE (m)", 500, 10000, 2000, step=100)

    st.markdown("**Fond de carte**")
    map_style_choice = st.selectbox(
        "Style",
        ["Clair (CARTO)", "Sombre (CARTO)", "Mapbox Light", "Satellite (ESRI – gratuit)", "Satellite (Mapbox)"],
        index=0
    )
    mapbox_token = None
    if "Mapbox" in map_style_choice:
        mapbox_token = st.text_input("Mapbox token (requis pour style Mapbox)", type="password")

    st.markdown("---")
    st.markdown("**Couche externe (KML/KMZ)**")
    kml_kmz_file = st.file_uploader("Charger une couche KML/KMZ", type=["kml", "kmz"])

    st.markdown("**Couche image (KMZ GroundOverlay)**")
    show_overlay = st.checkbox("Afficher l’image de couverture (si présente)", value=True)
    overlay_opacity = st.slider("Opacité de la couverture", 0.0, 1.0, 0.65, step=0.05)

    st.markdown("---")
    st.markdown("**Couleurs intelligentes** :")
    st.markdown("- Vert pâle : en-dessous du seuil")
    st.markdown("- Jaune → Orange → Rouge : ENU au-dessus du seuil (croissant)")

uploaded = st.file_uploader("Déposez un ou plusieurs rapports TXT", type=["txt"], accept_multiple_files=True)

df_sum = pd.DataFrame()
df_det = pd.DataFrame()

if uploaded:
    summaries = []; details_all = []; errors = []
    for up in uploaded:
        try:
            content = up.read().decode("utf-8", errors="ignore")
            summary, interfs = process_one_text(content, up.name, threshold)
            summaries.append(summary); details_all.extend(interfs)
        except Exception as e:
            errors.append((up.name, str(e)))

    df_sum = pd.DataFrame(summaries)
    order_sum = ['site','Canal','Fréquence','file','threshold_ENU_dB','risk','interferer_count','max_ENU','worst_transmitter','site_lon','site_lat']
    df_sum = df_sum[[c for c in order_sum if c in df_sum.columns] + [c for c in df_sum.columns if c not in order_sum]]
    df_sum = df_sum.sort_values(['risk','interferer_count','max_ENU'], ascending=[False, False, False])

    st.subheader("Résumé par site")
    st.dataframe(style_summary(df_sum, threshold), use_container_width=True)

    sites_opts = ["(Tous les sites)"] + sorted(df_sum['site'].dropna().unique().tolist())
    selected_site = st.selectbox("Carte/tableau : afficher le site…", sites_opts, index=0)

    if selected_site != "(Tous les sites)":
        canaux_disponibles = df_sum.loc[df_sum['site'] == selected_site, 'Canal'].dropna().unique().tolist()
    else:
        canaux_disponibles = df_sum['Canal'].dropna().unique().tolist()

    canaux_sorted = sorted(
        [int(x) if pd.notna(x) and str(x).isdigit() else x for x in canaux_disponibles],
        key=lambda z: (isinstance(z, str), z)
    )
    canaux_opts = ["(Tous les canaux)"] + canaux_sorted
    selected_canal = st.selectbox("… et le canal :", canaux_opts, index=0)

    st.subheader("Interférences (filtrées par sélection)")
    df_det = pd.DataFrame(details_all)
    if not df_det.empty:
        preferred = ['site','Canal','Fréquence','file','ENU','OS','TRANSMITTER','DIS','AZM','LONGITUDE','LATITUDE','ERP','f/MHz','CHA','HEFF','POL','PROGRAM','REMARKS','_ENU_float']
        cols = [c for c in preferred if c in df_det.columns] + [c for c in df_det.columns if c not in preferred]
        df_det = df_det[cols]

        df_det_display = df_det.copy()
        if selected_site != "(Tous les sites)":
            df_det_display = df_det_display[df_det_display['site'] == selected_site]
        if selected_canal != "(Tous les canaux)":
            df_det_display = df_det_display[df_det_display['Canal'] == selected_canal]

        if not df_det_display.empty:
            st.dataframe(style_details(df_det_display, threshold), use_container_width=True)
        else:
            st.info("Aucune interférence pour le filtre sélectionné.")
    else:
        st.info("Aucune interférence au-dessus du seuil.")

    # ====== Carte (avec filtre Site + Canal) ======
    df_det_view = df_det.copy() if not df_det.empty else pd.DataFrame()
    df_sum_view = df_sum.copy() if not df_sum.empty else pd.DataFrame()
    if selected_site != "(Tous les sites)":
        if not df_det_view.empty: df_det_view = df_det_view[df_det_view['site'] == selected_site]
        if not df_sum_view.empty: df_sum_view = df_sum_view[df_sum_view['site'] == selected_site]
    if selected_canal != "(Tous les canaux)":
        if not df_det_view.empty: df_det_view = df_det_view[df_det_view['Canal'] == selected_canal]
        if not df_sum_view.empty: df_sum_view = df_sum_view[df_sum_view['Canal'] == selected_canal]

    df_emitters = build_map_dataframe(df=df_det_view if not df_det_view.empty else pd.DataFrame(),
                                      threshold=threshold, only_interferers=only_map_interferers,
                                      marker_base=marker_base, marker_scale=marker_scale)
    df_sites = build_sites_dataframe(df_sum_view)

    # Préparer GeoJSON et/ou GroundOverlay depuis KML/KMZ si fourni  ✅ (bien indenté)
    geojson_data = None
    ground_overlays = []
    if kml_kmz_file is not None:
        try:
            kml_bytes = kml_kmz_file.getvalue()
            geojson_data = kml_kmz_to_geojson(kml_kmz_file.name, kml_bytes)
            ground_overlays = kmz_extract_groundoverlays(kml_kmz_file.name, kml_bytes)
            st.caption(
                f"Couche importée : "
                f"{len(geojson_data.get('features', [])) if geojson_data else 0} entité(s) vectorielle(s), "
                f"{len(ground_overlays)} image(s) de couverture."
            )
        except Exception as e:
            st.error(f"Échec de lecture KML/KMZ : {e}")
            geojson_data = None; ground_overlays = []

    # 👉 Appel rendu carte (toujours exécuté)
    render_map(
        df_emitters, df_sites,
        title="Carte des émetteurs et sites étudiés (filtre Site+Canal)",
        map_style_choice=map_style_choice,
        mapbox_token=mapbox_token,
        site_marker_radius=site_marker_radius,
        geojson_overlay=geojson_data,
        ground_overlays=ground_overlays,
        overlay_opacity=overlay_opacity,
        show_overlay=show_overlay
    )

    # Exports
    if not df_sum.empty:
        xls_bytes = export_to_excel(df_sum, df_det)
        st.download_button("⬇️ Télécharger l'Excel coloré", data=xls_bytes, file_name="resultats_interferences.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        pdf_bytes = export_to_pdf(df_sum, df_det)
        st.download_button("⬇️ Télécharger le rapport PDF", data=pdf_bytes, file_name="rapport_interferences.pdf",
                           mime="application/pdf")

    if errors:
        st.warning("Certains fichiers ont été traités avec avertissements.")
        for name, msg in errors:
            st.write(f"• **{name}** : {msg}")
else:
    st.info("Chargez vos fichiers TXT pour lancer l'analyse.")
